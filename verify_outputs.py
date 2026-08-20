# -*- coding: utf-8 -*-
"""端到端验证：检查计算过程表、日报Word、汇总Excel 的关键内容。"""
import pandas as pd
from openpyxl import load_workbook

PROC = "测试输出/BI_ADI_计算过程_08月20日.xlsx"
WORD = "测试输出/省媒介伊蚊传染病疫情蚊媒监测情况（08月20日 20：00）.docx"
SUM = "测试输出/全省媒介伊蚊传染病疫点重点镇（街道）蚊媒密度监测村居一览表（08月20日 20：00）.xlsx"

print("=" * 70)
print("一、计算过程表")
wb = load_workbook(PROC)
print("工作表：", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row - 1} 行数据")

ws1 = wb["BI+SSI表"]
print("\n[BI+SSI表] 塘西村(大岗镇)两条记录：")
for r in range(2, ws1.max_row + 1):
    region = ws1.cell(r, 1).value
    village = ws1.cell(r, 2).value
    if village == "塘西村" and region and "大岗镇" in str(region):
        print(f"  方法={ws1.cell(r,7).value} 值={ws1.cell(r,8).value} 地址={ws1.cell(r,3).value}")

ws2 = wb["BI+SSI(不重复)"]
yellow = []
for r in range(2, ws2.max_row + 1):
    fill = ws2.cell(r, 1).fill
    if fill and fill.fgColor and fill.fgColor.rgb and "FFEB9C" in str(fill.fgColor.rgb):
        yellow.append(ws2.cell(r, 2).value)
print(f"\n[BI+SSI(不重复)] 黄色(SSI来源)行数={len(yellow)} 村居={yellow}")

ws3 = wb["BI+SSI(重复)+取较大值处理"]
print("\n[BI+SSI(重复)+取较大值处理]")
for r in range(2, ws3.max_row + 1):
    print("  值=%s 原BI值=%s 原SSI值=%s 转换后SSI值=%s 村居=%s" % (
        ws3.cell(r, 8).value, ws3.cell(r, 9).value, ws3.cell(r, 10).value,
        ws3.cell(r, 11).value, ws3.cell(r, 2).value))

ws4 = wb["重复数据删除"]
red = [ws4.cell(r, 2).value for r in range(2, ws4.max_row + 1)
       if ws4.cell(r, 1).fill and ws4.cell(r, 1).fill.fgColor
       and str(ws4.cell(r, 1).fill.fgColor.rgb).endswith("FF7C80")]
print(f"\n[重复数据删除] 标红(被删除)行数={len(red)} 村居={red}")

ws5 = wb["地址区分处理"]
print("\n[地址区分处理]")
for r in range(2, ws5.max_row + 1):
    if ws5.cell(r, 9).value:  # 备注非空
        print(f"  监测地点={ws5.cell(r,6).value} | {ws5.cell(r,9).value}")

ws6 = wb["最终表"]
methods = {ws6.cell(r, 7).value for r in range(2, ws6.max_row + 1)}
print(f"\n[最终表] 方法集合={methods} 行数={ws6.max_row - 1}")

wsa = wb["ADI(重复)+取较大值处理"]
print("\n[ADI(重复)+取较大值处理]")
for r in range(2, wsa.max_row + 1):
    print(f"  值={wsa.cell(r,8).value} 原ADI值={wsa.cell(r,9).value} 村居={wsa.cell(r,2).value}")

print("=" * 70)
print("二、日报Word")
from docx import Document
doc = Document(WORD)
for p in doc.paragraphs:
    if p.text.strip():
        style = p.style.name if p.style else ""
        print(f"[{style}] {p.text}")

print("=" * 70)
print("三、监测点汇总表")
sdf_bi = pd.read_excel(SUM, sheet_name="BI表")
print("BI表前8行：")
print(sdf_bi.head(8).to_string(index=False))
sdf_adi = pd.read_excel(SUM, sheet_name="ADI表")
print("\nADI表：")
print(sdf_adi.to_string(index=False))
del_df = pd.read_excel(SUM, sheet_name="删除数据情况说明")
print(f"\n删除数据情况说明：{len(del_df)} 行")
print(del_df[["地市-区/县/市-街道/乡/镇", "社区/村居", "监测指标值", "删除原因"]].to_string(index=False))
