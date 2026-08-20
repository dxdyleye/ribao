# -*- coding: utf-8 -*-
"""Excel 输出：计算过程表（BI_ADI_计算过程_MM月DD日.xlsx）与监测点汇总表。"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config as C
from . import utils as U
from .processor import ADI_SHEET_ORDER, BI_SHEET_ORDER


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _to_excel_table(df):
    """去掉内部辅助列，得到真正写入 Excel 的表。"""
    drop = [c for c in (C.ID_COL, C.KEY_COL, C.KIND_COL, C.ORIG_COL, C.FILL_COL) if c in df.columns]
    return df.drop(columns=drop)


def _style_sheet(ws, df, fill_series=None, number_cols=(), center_cols=()):
    """表头加粗、填充、冻结、列宽、数值格式等基础样式。"""
    ncols = ws.max_column
    fill_series = fill_series if fill_series is not None else []

    # 数据行填充（黄色/红色整行标记）
    for i, f in enumerate(fill_series):
        if not f:
            continue
        color = C.FILL_YELLOW if f == "yellow" else C.FILL_RED
        for col in range(1, ncols + 1):
            ws.cell(row=i + 2, column=col).fill = _fill(color)

    # 表头
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = _fill(C.FILL_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 数值格式
    col_index = {name: i + 1 for i, name in enumerate(df.columns)}
    for cname in number_cols:
        ci = col_index.get(cname)
        if ci:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=ci).number_format = "0.0"
    for cname in center_cols:
        ci = col_index.get(cname)
        if ci:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=ci).alignment = Alignment(horizontal="center")

    # 列宽
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        width = 8
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                width = max(width, min(50, len(str(v))))
        ws.column_dimensions[letter].width = width + 2

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


# ---------------------------------------------------------------------------
# 1. 计算过程表
# ---------------------------------------------------------------------------
def build_process_excel(result, outpath):
    """生成 BI_ADI_计算过程_MM月DD日.xlsx（含全部 BI/ADI 中间表）。"""
    sheets = {}
    for name in BI_SHEET_ORDER:
        sheets[name] = result.bi_sheets[name]
    for name in ADI_SHEET_ORDER:
        sheets[name] = result.adi_sheets[name]

    with pd.ExcelWriter(outpath, engine="openpyxl") as writer:
        for name, df in sheets.items():
            _to_excel_table(df).to_excel(writer, sheet_name=name, index=False)

    wb = load_workbook(outpath)
    for name, df in sheets.items():
        ws = wb[name]
        fill_series = df[C.FILL_COL].tolist() if C.FILL_COL in df.columns else [None] * len(df)
        _style_sheet(
            ws, _to_excel_table(df), fill_series=fill_series,
            number_cols=("监测指标值", "原BI值", "原SSI值", "转换后的SSI值"),
        )
    wb.save(outpath)
    return outpath


# ---------------------------------------------------------------------------
# 2. 监测点汇总表
# ---------------------------------------------------------------------------
def _summary_table(final, label):
    """最终表 -> 汇总表（地市/区县/街道/监测地点/BI*|ADI*/风险水平*）。"""
    recs = []
    for _, r in final.iterrows():
        city, district, street = U.parse_region(r[C.COL_REGION])
        value = float(r[C.COL_VALUE])
        recs.append({
            "地市": U.city_short(city),
            "区县": U.district_short(district),
            "街道": street,
            "监测地点": r["监测地点"],
            f"{label}*": round(value, 1),
            "风险水平*": U.risk_level(value),
            "_zone": r[C.COL_ZONE],  # 仅用于排序
        })
    df = pd.DataFrame(recs)
    if df.empty:
        return pd.DataFrame(columns=["地市", "区县", "街道", "监测地点", f"{label}*", "风险水平*"])
    known = list(C.REGION_ORDER)
    extra = sorted(set(df["地市"]) - set(known))
    df["地市"] = pd.Categorical(df["地市"], categories=known + extra, ordered=True)
    df = df.sort_values(["地市", "区县", "街道", "监测地点", "_zone"], kind="stable")
    df["地市"] = df["地市"].astype(str)
    return df.drop(columns=["_zone"])


_RISK_COLORS = {
    "安全": C.FILL_GREEN,
    "低风险": C.FILL_YELLOW,
    "中风险": C.FILL_ORANGE,
    "高风险": C.FILL_RED,
}


def build_summary_excel(result, outpath):
    """生成监测点汇总 Excel（BI表 / ADI表 / 删除数据情况说明）。"""
    bi_df = _summary_table(result.bi_final, "BI")
    adi_df = _summary_table(result.adi_final, "ADI")
    del_df = result.deletions

    with pd.ExcelWriter(outpath, engine="openpyxl") as writer:
        bi_df.to_excel(writer, sheet_name="BI表", index=False)
        adi_df.to_excel(writer, sheet_name="ADI表", index=False)
        del_df.to_excel(writer, sheet_name="删除数据情况说明", index=False)

    wb = load_workbook(outpath)

    for name, df, label in (("BI表", bi_df, "BI"), ("ADI表", adi_df, "ADI")):
        ws = wb[name]
        col_index = {cname: i + 1 for i, cname in enumerate(df.columns)}
        # 风险水平单元格按等级填充背景色
        ci = col_index["风险水平*"]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=ci).value
            if v in _RISK_COLORS:
                ws.cell(row=r, column=ci).fill = _fill(_RISK_COLORS[v])
        _style_sheet(ws, df, number_cols=(f"{label}*",), center_cols=("风险水平*",))

    ws3 = wb["删除数据情况说明"]
    _style_sheet(ws3, del_df)

    wb.save(outpath)
    return outpath


# ---------------------------------------------------------------------------
# 3. 输出文件路径
# ---------------------------------------------------------------------------
def process_excel_path(outdir, target):
    mm, dd = f"{target.month:02d}", f"{target.day:02d}"
    return f"{outdir}/BI_ADI_计算过程_{mm}月{dd}日.xlsx"


def word_path(outdir, target):
    mm, dd = f"{target.month:02d}", f"{target.day:02d}"
    return f"{outdir}/省媒介伊蚊传染病疫情蚊媒监测情况（{mm}月{dd}日 20：00）.docx"


def summary_excel_path(outdir, target):
    mm, dd = f"{target.month:02d}", f"{target.day:02d}"
    return f"{outdir}/全省媒介伊蚊传染病疫点重点镇（街道）蚊媒密度监测村居一览表（{mm}月{dd}日 20：00）.xlsx"
